from django.shortcuts import render , redirect
from django.db.models import Count
from .models import InventoryItem, ProductCategory
from django.contrib import messages
from django.db import models
from django.core.management import call_command
from django.http import FileResponse
from django.contrib.staticfiles import finders
from .utils import process_inventory_file 
import pandas as pd
from django.contrib.auth.decorators import login_required 
from django.db import transaction
from django.http import JsonResponse
from django.db.models import Q
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter

@login_required
def inventory_list_view(request):
    search_query = request.GET.get('q', '')

    # Start with the base query that we will group later
    base_queryset = InventoryItem.objects.all()

    # If there is a search query, we apply the filter logic
    if search_query:
        # We want to find items where EITHER the serial number contains the query
        # OR the related category's name contains the query.
        # This gives us a list of all individual items that match.
        matching_items = base_queryset.filter(
            models.Q(serial_number__icontains=search_query) |
            models.Q(category__name__icontains=search_query)
        )
        
        # Now, we find out which unique category IDs these matching items belong to.
        matching_category_ids = matching_items.values_list('category_id', flat=True).distinct()
        
        # We will now use these category IDs to filter our main summary query.
        # However, the summary query is on the base_queryset, so we can just filter it directly.
        base_queryset = base_queryset.filter(category_id__in=matching_category_ids)

    # Now, perform the aggregation on the (potentially filtered) base_queryset
    summary_query = base_queryset.values(
        'location', 
        'category__id', 
        'category__name'
    ).annotate(
        quantity=Count('id')
    ).order_by('location', 'category__name')

    # The rest of the logic remains the same
    maadi_summary = []
    abu_rudies_summary = []

    for item in summary_query:
        location = item.get('location', '').strip()
        if location == 'maadi-yard':
            maadi_summary.append(item)
        elif location == 'abu-rudies-yard':
            abu_rudies_summary.append(item)
    
    context = {
        'maadi_summary': maadi_summary,
        'abu_rudies_summary': abu_rudies_summary,
        'search_query': search_query,
    }
    return render(request, 'inventory/inventory_list.html', context)

@login_required
def get_item_details_view(request, category_id, location):
    # This view is now simple again. It only fetches and returns details.
    items = InventoryItem.objects.filter(category_id=category_id, location=location)
    context = {
        'items': items
    }
    return render(request, 'inventory/_item_details.html', context)

@login_required
def inventory_filtered_list_view(request):
    title = "Inventory Details"
    status_filter = request.GET.get('status', None)
    search_query = request.GET.get('q', '')

    queryset = InventoryItem.objects.all()

    if status_filter:
        if status_filter == 'available':
            title = "Available Items"
            queryset = queryset.filter(status='available')
        elif status_filter == 'on_job':
            title = "Items On Job"
            queryset = queryset.filter(status='on_job')
        elif status_filter == 'pending_inspection':
            title = "Items Pending Inspection"
            queryset = queryset.filter(status='pending_inspection')
        elif status_filter == 're-cut':
            title = "Items to be Recut"
            queryset = queryset.filter(status='re-cut')
        elif status_filter == 'lih':
            title = "LIH Items"
            queryset = queryset.filter(status='lih')
        elif status_filter == 'junk':
            title = "Junk Items"
            queryset = queryset.filter(status='junk')
        elif status_filter == 'sold':
            title = "Sold Items"
            queryset = queryset.filter(status='sold')

    if search_query:
        queryset = queryset.filter(
            models.Q(serial_number__icontains=search_query) |
            models.Q(category__name__icontains=search_query)
        )

    context = {
        'title': title,
        'items': queryset,
        'search_query': search_query,
        'status_filter': status_filter,
    }
    return render(request, 'inventory/inventory_filtered_list.html', context)

@login_required
def download_template_view(request):
    file_path = finders.find('downloads/import_template.xlsx')
    response = FileResponse(open(file_path, 'rb'), as_attachment=True, filename='import_template.xlsx')
    return response

@login_required
def import_results_view(request):
    summary = request.session.get('import_summary', None)
    # Clear the summary from the session so it doesn't show again on refresh
    if 'import_summary' in request.session:
        del request.session['import_summary']
    
    return render(request, 'inventory/import_results.html', {'summary': summary})

@login_required
def inventory_import_view(request):
    if request.method == 'POST':
        file = request.FILES.get('inventory_file')

        if not file:
            messages.error(request, "No file was selected for upload.")
            return redirect('inventory_import')
        
        try:
            # Read the file into a pandas DataFrame
            if file.name.endswith('.xlsx'):
                df = pd.read_excel(file)
            elif file.name.endswith('.csv'):
                df = pd.read_csv(file)
            else:
                messages.error(request, "Unsupported file format. Please upload a .xlsx or .csv file.")
                return redirect('inventory_import')

            # Call our reusable processing function
            summary = process_inventory_file(df, strict=False)

            # Recalculate quantities if there were no critical errors
            if (summary.get("created", 0) + summary.get("updated", 0)) > 0:
              call_command('recalculate_quantities')

            
            # Store summary in session and redirect to results page
            request.session['import_summary'] = summary
            return redirect('import_results')

        except Exception as e:
            messages.error(request, f"An error occurred while processing the file: {e}")
            return redirect('inventory_import')

    return render(request, 'inventory/import_form.html')

@login_required
def inventory_change_status_view(request):
    if request.method == "POST":
        selected_ids = request.POST.getlist("selected_items")
        new_status = request.POST.get("new_status", "").strip()

        if not selected_ids:
            messages.error(request, "Please select at least one item.")
            return redirect("inventory_change_status")

        valid_statuses = {k for k, _ in InventoryItem.STATUS_CHOICES}
        if new_status not in valid_statuses:
            messages.error(request, "Invalid status selected.")
            return redirect("inventory_change_status")

        with transaction.atomic():
            items = list(InventoryItem.objects.select_for_update().filter(id__in=selected_ids))
            InventoryItem.objects.filter(id__in=[i.id for i in items]).update(status=new_status)

        serials = list(InventoryItem.objects.filter(id__in=selected_ids).values_list('serial_number', flat=True))

        messages.success(request,f"Status updated to '{new_status}' for: {', '.join(serials)}")
        return redirect("inventory_change_status")

    return render(request, "inventory/change_status.html", {
        "status_choices": InventoryItem.STATUS_CHOICES
    })

@login_required
def ajax_inventory_search(request):
    q = (request.GET.get("q") or "").strip()
    if not q:
        return JsonResponse([], safe=False)

    qs = InventoryItem.objects.select_related("category").filter(
        Q(serial_number__icontains=q) | Q(category__name__icontains=q)
    )[:30]

    data = []
    for item in qs:
        data.append({
            "id": item.id,
            "serial": item.serial_number,
            "category": item.category.name,
            "location": item.get_location_display(),
            "status": item.get_status_display(),
        })

    return JsonResponse(data, safe=False)

@login_required
def export_inventory_to_excel_view(request):
    """
    This view handles the export of filtered inventory items to an Excel file.
    """
    status_filter = request.GET.get('status', None)
    
    # --- Re-use the exact same filtering logic from inventory_filtered_list_view ---
    queryset = InventoryItem.objects.all().select_related('category')
    title = "Full Inventory"

    if status_filter:
        if status_filter == 'available':
            title = "Available_Items"
            queryset = queryset.filter(status='available')
        elif status_filter == 'on_job':
            title = "On_Job_Items"
            queryset = queryset.filter(status='on_job')
        elif status_filter == 'pending_inspection':
            title = "Pending_Inspection_Items"
            queryset = queryset.filter(status='pending_inspection')
        elif status_filter == 're-cut':
            title = "Recut_Items"
            queryset = queryset.filter(status='re-cut')
        elif status_filter == 'lih':
            title = "LIH_Items"
            queryset = queryset.filter(status='lih')
        elif status_filter == 'junk':
            title = "Junk_Items"
            queryset = queryset.filter(status='junk')
        elif status_filter == 'sold':
            title = "Sold_Items"
            queryset = queryset.filter(status='sold')

    # --- Create the Excel file in memory ---
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = title

    # Define headers
    headers = ['Serial Number', 'Category', 'Location', 'Status', 'Updated At']
    if status_filter == 're-cut':
        headers.append('Reason') # Add the 'Reason' header only for re-cut items

    # Write headers to the first row
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        # Auto-adjust column width
        sheet.column_dimensions[get_column_letter(col_num)].width = 20

    # Write data rows
    for row_num, item in enumerate(queryset, 2):
        sheet.cell(row=row_num, column=1, value=item.serial_number)
        sheet.cell(row=row_num, column=2, value=item.category.name)
        sheet.cell(row=row_num, column=3, value=item.get_location_display())
        sheet.cell(row=row_num, column=4, value=item.get_status_display())
        sheet.cell(row=row_num, column=5, value=item.updated_at.strftime('%Y-%m-%d %H:%M'))
        
        if status_filter == 're-cut':
            sheet.cell(row=row_num, column=6, value=item.recut_reason)

    # --- Prepare the HTTP response ---
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    # Set the file name for the download
    response['Content-Disposition'] = f'attachment; filename="{title}.xlsx"'
    
    # Save the workbook to the response
    workbook.save(response)

    return response
